#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ETL - Observatorio de Mercado Laboral | Camara de Comercio de Cali
==================================================================

Lee los anexos mensuales de la GEIH (DANE) que esten en ./datos y produce
./docs/datos.json, que es lo unico que consume el tablero.

Uso:
    python etl.py                 # busca los anexos mas recientes en ./datos
    python etl.py --datos ruta/   # otra carpeta de insumos

Los archivos se identifican por PATRON de nombre, no por nombre exacto, para
que el proceso siga funcionando cuando el DANE cambie el sufijo del mes:

    anexGEIH<mes><anio>.xlsx          -> modulo general      (patron: sin ML/EISS)
    anexGEIHMLS<periodo>.xlsx         -> modulo sexo         (MLS)
    anexGEIHEISS<periodo>.xlsx        -> modulo informalidad (EISS)
    anexGEIHMLJ<periodo>.xlsx         -> modulo juventud     (MLJ)

Las filas y columnas NO estan quemadas: los bloques de ciudad y los
indicadores se localizan por su etiqueta de texto en la columna A.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------

RAIZ = Path(__file__).resolve().parent

# Ancla de las series de trimestre movil: la primera columna de datos es el
# trimestre que TERMINA en marzo de 2007. La columna i termina i meses despues.
ANCLA_TM = (2007, 3)
ANCLA_INFORMALIDAD = (2021, 3)

MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
         "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

CIUDAD_FOCO = "Cali A.M."

# Los anexos escriben las ciudades en mayusculas, con y sin tilde. Se
# normalizan a una clave sin tildes y luego se reponen para mostrarlas bien.
NOMBRES_BONITOS = {
    "Bogota D.C.": "Bogotá D.C.", "Medellin A.M.": "Medellín A.M.",
    "Cucuta A.M.": "Cúcuta A.M.", "Ibague": "Ibagué", "Monteria": "Montería",
    "Popayan": "Popayán", "Quibdo": "Quibdó",
}

# Orden en que se muestran las ciudades en el tablero
ORDEN_CIUDADES = [
    "Total nacional", "Total 13 ciudades y A.M.", "Total 23 ciudades y A.M.",
    "Bogota D.C.", "Medellin A.M.", "Cali A.M.",
    "Barranquilla A.M.", "Bucaramanga A.M.", "Cartagena", "Cucuta A.M.",
    "Pereira A.M.", "Manizales A.M.", "Ibague", "Monteria", "Villavicencio",
    "Pasto", "Neiva", "Armenia", "Popayan", "Santa Marta", "Valledupar",
    "Sincelejo", "Riohacha", "Tunja", "Florencia", "Quibdo",
]


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def norm(txt) -> str:
    """Minusculas, sin tildes, sin espacios repetidos. Para comparar etiquetas."""
    if txt is None:
        return ""
    s = str(txt).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s.lower()


def titulo_ciudad(txt) -> str:
    """'CALI A.M.' / 'Cali A.M.' -> 'Cali A.M.' (sin tildes, forma canonica)."""
    s = str(txt).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    # Title case respetando siglas A.M. y D.C.
    partes = []
    for w in s.split(" "):
        if w.upper() in ("A.M.", "D.C.", "A.M", "D.C"):
            partes.append(w.upper())
        else:
            partes.append(w.capitalize())
    s = " ".join(partes)
    s = s.replace("Total 23 Ciudades Y Areas Metropolitanas", "Total 23 ciudades y A.M.")
    s = s.replace("Total 13 Ciudades Y Areas Metropolitanas", "Total 13 ciudades y A.M.")
    s = s.replace("23 Ciudades Y A.M.", "Total 23 ciudades y A.M.")
    s = s.replace("13 Ciudades Y A.M.", "Total 13 ciudades y A.M.")
    s = s.replace("13 Ciudades Y A.M", "Total 13 ciudades y A.M.")
    return s


def num(v):
    """
    Convierte a float. Un cero exacto NO es un dato: en estos anexos el DANE
    rellena con 0 lo que no midio. El caso claro es la subocupacion entre
    Ene-Mar y Jul-Sep de 2020, cuando la pandemia interrumpio la recoleccion.
    En un area metropolitana de millones de habitantes ninguna poblacion ni
    tasa de este tablero puede valer exactamente cero, asi que se trata como
    dato faltante y la grafica deja el hueco en vez de dibujar una caida.
    """
    if v is None or isinstance(v, str):
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if f == 0:
        return None
    return round(f, 2) if abs(f) < 200 else round(f, 1)


def cod_periodo(anio: int, mes_fin: int) -> str:
    """Codigo estable del trimestre movil: '2026-06' = trimestre Abr-Jun 2026."""
    return f"{anio}-{mes_fin:02d}"


def etiqueta_tm(anio: int, mes_fin: int) -> str:
    """'Abr - Jun 2026'."""
    ini = (mes_fin - 3) % 12 + 1
    anio_ini = anio if mes_fin >= 3 else anio - 1
    if anio_ini == anio:
        return f"{MESES[ini - 1]} - {MESES[mes_fin - 1]} {anio}"
    return f"{MESES[ini - 1]} {str(anio_ini)[2:]} - {MESES[mes_fin - 1]} {str(anio)[2:]}"


def etiqueta_anual(anio: int, n_trim: int) -> str:
    """'2025' si el ano esta completo; '2026 (Ene-Jun)' si va a medias."""
    if n_trim >= 4:
        return str(anio)
    return f"{anio} (Ene-{MESES[n_trim * 3 - 1]})"


def periodos_desde(ancla, n: int):
    """Lista de (codigo, etiqueta, anio, mes_fin) para n periodos consecutivos."""
    a0, m0 = ancla
    out = []
    for i in range(n):
        t = a0 * 12 + (m0 - 1) + i
        anio, mes = divmod(t, 12)
        mes += 1
        out.append((cod_periodo(anio, mes), etiqueta_tm(anio, mes), anio, mes))
    return out


# ---------------------------------------------------------------------------
# Localizacion de archivos
# ---------------------------------------------------------------------------

def localizar_anexos(carpeta: Path) -> dict:
    """Devuelve {'general':Path,'sexo':Path,'informalidad':Path,'juventud':Path}."""
    archivos = sorted(carpeta.glob("*.xlsx"))
    if not archivos:
        sys.exit(f"[ERROR] No hay archivos .xlsx en {carpeta}")

    encontrados = {}
    for f in archivos:
        if f.name.startswith("~$"):
            continue
        # El DANE sirve "anex-GEIH-jun2026.xlsx" y el navegador a veces guarda
        # "anexGEIHjun2026 (2).xlsx". Se quitan guiones, espacios y parentesis
        # para que ambos nombres lleguen al mismo lugar.
        n = re.sub(r"[^a-z0-9]", "", f.name.lower())
        # El DANE publica junto al anexo general otros que NO usa este tablero
        if any(x in n for x in ("desestacionalizado", "relab", "creativa")):
            continue
        if "eiss" in n:
            encontrados.setdefault("informalidad", []).append(f)
        elif "mlj" in n:
            encontrados.setdefault("juventud", []).append(f)
        elif "mls" in n:
            encontrados.setdefault("sexo", []).append(f)
        elif n.startswith("anexgeih"):
            encontrados.setdefault("general", []).append(f)

    res = {}
    for k, v in encontrados.items():
        # el mas reciente por fecha de modificacion
        res[k] = max(v, key=lambda p: p.stat().st_mtime)

    faltan = {"general", "sexo", "informalidad", "juventud"} - set(res)
    if faltan:
        print(f"[AVISO] No se encontro anexo para: {', '.join(sorted(faltan))}")
    for k, v in sorted(res.items()):
        print(f"  [ok] {k:<13} -> {v.name}")
    return res


def urls_de_origen(anexos: dict) -> dict:
    """De que URL del DANE salio cada anexo, para poder citarla en el tablero.

    Primero se mira datos/historial.csv, que deja descargar.py con la URL
    exacta de la que bajo cada archivo. Es la fuente autoritativa: dice de
    donde salio ese archivo, no de donde deberia haber salido.

    Si no hay historial -- porque los anexos entraron a mano, como paso hasta
    que se automatizo -- se reconstruye la URL canonica a partir del periodo
    que declara el nombre del archivo. Si ni eso se puede, queda vacia: es
    preferible no ofrecer un enlace a ofrecer uno roto.
    """
    from descargar import ARCHIVOS, MODULOS, cierre_de, nombre_esperado

    historial = {}
    ruta_hist = next(iter(anexos.values())).parent / "historial.csv"
    if ruta_hist.exists():
        for linea in ruta_hist.read_text(encoding="utf-8").splitlines()[1:]:
            campos = linea.split(";")
            if len(campos) >= 6:
                historial[campos[3]] = campos[5]

    salida = {}
    for modulo, ruta in anexos.items():
        url = historial.get(ruta.name, "")
        if not url:
            cierre = cierre_de(ruta.name)
            if cierre and modulo in MODULOS:
                url = f"{ARCHIVOS}/{nombre_esperado(modulo, *cierre)}"
        salida[modulo] = {
            "archivo": ruta.name,
            "url": url,
            "pagina": MODULOS.get(modulo, {}).get("pagina", ""),
        }
    return salida


# ---------------------------------------------------------------------------
# Parser generico de hojas por bloques de ciudad
# ---------------------------------------------------------------------------

def cargar_hoja(ruta: Path, hoja: str):
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    if hoja not in wb.sheetnames:
        # tolerar espacios sobrantes en el nombre de hoja
        cand = [s for s in wb.sheetnames if norm(s) == norm(hoja)]
        if not cand:
            wb.close()
            raise KeyError(f"Hoja '{hoja}' no existe en {ruta.name}. Hay: {wb.sheetnames}")
        hoja = cand[0]
    ws = wb[hoja]
    filas = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return filas


def detectar_bloques(filas, nombres_validos):
    """Devuelve [(fila_indice, nombre_canonico)] de cada bloque de ciudad."""
    bloques = []
    validos = {norm(n) for n in nombres_validos}
    for i, fila in enumerate(filas):
        if not fila:
            continue
        v = fila[0]
        if v is None:
            continue
        nv = norm(v)
        if nv in validos:
            bloques.append((i, titulo_ciudad(v)))
    return bloques


def indicadores_del_bloque(filas, fila_ini, fila_fin, mapa, n_periodos, col_ini=1):
    """
    Extrae series del bloque [fila_ini, fila_fin).
    mapa: {clave_salida: prefijo_etiqueta_normalizado}
    """
    out = {}
    for i in range(fila_ini, min(fila_fin, len(filas))):
        fila = filas[i]
        if not fila or fila[0] is None:
            continue
        etq = norm(fila[0])
        for clave, prefijo in mapa.items():
            if clave in out:
                continue
            if etq.startswith(prefijo):
                serie = []
                for c in range(col_ini, col_ini + n_periodos):
                    serie.append(num(fila[c]) if c < len(fila) else None)
                out[clave] = serie
    return out


def contar_periodos(filas, fila_encabezado, col_ini=1):
    """Cuenta columnas con etiqueta de trimestre en la fila de encabezado."""
    fila = filas[fila_encabezado]
    n = 0
    for c in range(col_ini, len(fila)):
        if fila[c] is None or str(fila[c]).strip() == "":
            break
        n += 1
    return n


# ---------------------------------------------------------------------------
# Modulo GENERAL (23 ciudades, trimestre movil 2007-)
# ---------------------------------------------------------------------------

MAPA_GENERAL = {
    "pct_pet": "% poblacion en edad de trabajar",
    "tgp":     "tasa global de participacion",
    "to":      "tasa de ocupacion",
    "td":      "tasa de desocupacion",
    "ts":      "tasa de subocupacion",
    "pob_total": "poblacion total",
    "pet":     "poblacion en edad de trabajar",
    "ft":      "fuerza de trabajo",
    "ocupados": "poblacion ocupada",
    "desocupados": "poblacion desocupada",
    "ffft":    "poblacion fuera de la fuerza de trabajo",
    "subocupados": "poblacion subocupada",
}

NOMBRES_CIUDAD = [
    "Total 13 ciudades y áreas metropolitanas",
    "Total 23 ciudades y áreas metropolitanas",
    "Bogotá D.C.", "Medellín A.M.", "Cali A.M.", "Barranquilla A.M.",
    "Bucaramanga A.M.", "Manizales A.M.", "Pasto", "Pereira A.M.",
    "Cúcuta A.M.", "Ibagué", "Montería", "Cartagena", "Villavicencio",
    "Tunja", "Florencia", "Popayán", "Valledupar", "Quibdó", "Neiva",
    "Riohacha", "Santa Marta", "Armenia", "Sincelejo",
]


def parse_general(ruta: Path):
    filas = cargar_hoja(ruta, "Total 23 ciudades A.M. Trim")
    bloques = detectar_bloques(filas, NOMBRES_CIUDAD)
    if not bloques:
        sys.exit("[ERROR] No se detectaron bloques de ciudad en el modulo general.")

    # fila de encabezado de trimestres = fila del bloque + 3
    n_per = contar_periodos(filas, bloques[0][0] + 3)
    periodos = periodos_desde(ANCLA_TM, n_per)

    datos = {}
    for k, (fi, nombre) in enumerate(bloques):
        ff = bloques[k + 1][0] if k + 1 < len(bloques) else len(filas)
        datos[nombre] = indicadores_del_bloque(filas, fi, ff, MAPA_GENERAL, n_per)

    print(f"  general      : {len(datos)} ciudades x {n_per} trimestres moviles "
          f"({periodos[0][1]} -> {periodos[-1][1]})")
    return periodos, datos


# ---------------------------------------------------------------------------
# Modulo SEXO (23 ciudades, trimestre movil 2007-)
# ---------------------------------------------------------------------------

MAPA_SEXO = {
    "pct_pet": "% poblacion en edad de trabajar",
    "tgp": "tasa global de participacion",
    "to":  "tasa de ocupacion",
    "td":  "tasa de desocupacion",
    "pob_total": "poblacion total",
    "pet": "poblacion en edad de trabajar",
    "ft":  "fuerza de trabajo",
    "ocupados": "poblacion ocupada",
    "desocupados": "poblacion desocupada",
}


def parse_total_nacional(ruta: Path, n_per: int):
    """La serie nacional vive en su propia hoja, sin bloques de ciudad."""
    try:
        filas = cargar_hoja(ruta, "Total nacional Trim")
    except KeyError:
        return None
    fila_hdr = next((i for i, f in enumerate(filas)
                     if f and len(f) > 1 and f[1] and norm(f[1]).startswith("ene - mar")), None)
    if fila_hdr is None:
        return None
    return indicadores_del_bloque(filas, fila_hdr, len(filas), MAPA_GENERAL, n_per)


def parse_sexo(ruta: Path):
    salida = {}
    n_per = None
    periodos = None
    for hoja, clave in (("Hombres - 23 Ciud", "hombres"), ("Mujeres - 23 Ciud", "mujeres")):
        filas = cargar_hoja(ruta, hoja)
        bloques = detectar_bloques(filas, NOMBRES_CIUDAD)
        if n_per is None:
            n_per = contar_periodos(filas, bloques[0][0] + 2)
            periodos = periodos_desde(ANCLA_TM, n_per)
        d = {}
        for k, (fi, nombre) in enumerate(bloques):
            ff = bloques[k + 1][0] if k + 1 < len(bloques) else len(filas)
            d[nombre] = indicadores_del_bloque(filas, fi, ff, MAPA_SEXO, n_per)
        salida[clave] = d
    print(f"  sexo         : {len(salida['mujeres'])} ciudades x {n_per} trimestres moviles")
    return periodos, salida


# ---------------------------------------------------------------------------
# Modulo JUVENTUD (15-28 anios, 23 ciudades, trimestre movil 2007-)
# ---------------------------------------------------------------------------

MAPA_JOVEN = {
    "pct_pet_joven": "% poblacion en edad de trabajar",
    "tgp": "tasa global de participacion",
    "to":  "tasa de ocupacion",
    "td":  "tasa de desocupacion",
    "pct_ffft": "% fuera de la fuerza de trabajo",
    "pet_total": "poblacion en edad de trabajar total",
    "ft":  "fuerza de trabajo",
    "ocupados": "poblacion ocupada",
    "desocupados": "poblacion desocupada",
    "ffft": "poblacion fuera de la fuerza",
}


def parse_juventud(ruta: Path):
    filas = cargar_hoja(ruta, "23 ciudades trim móvil")
    bloques = detectar_bloques(filas, NOMBRES_CIUDAD)
    n_per = contar_periodos(filas, bloques[0][0] + 2)
    periodos = periodos_desde(ANCLA_TM, n_per)
    datos = {}
    for k, (fi, nombre) in enumerate(bloques):
        ff = bloques[k + 1][0] if k + 1 < len(bloques) else len(filas)
        b = indicadores_del_bloque(filas, fi, ff, MAPA_JOVEN, n_per)
        # "Poblacion en edad de trabajar (15-28)" aparece 2 veces: total y joven.
        # La segunda ocurrencia es la joven -> la recuperamos aparte.
        ocur = [i for i in range(fi, min(ff, len(filas)))
                if filas[i] and filas[i][0] and
                norm(filas[i][0]).startswith("poblacion en edad de trabajar")]
        if len(ocur) >= 2:
            f2 = filas[ocur[1]]
            b["pet_joven"] = [num(f2[c]) if c < len(f2) else None
                              for c in range(1, 1 + n_per)]
        datos[nombre] = b
    print(f"  juventud     : {len(datos)} ciudades x {n_per} trimestres moviles")
    return periodos, datos


# ---------------------------------------------------------------------------
# Modulo INFORMALIDAD (23 ciudades, trimestre movil 2021-)
# ---------------------------------------------------------------------------

def parse_informalidad(ruta: Path):
    # 1) proporcion de informalidad por ciudad
    filas = cargar_hoja(ruta, "Prop informalidad")
    fila_hdr = next(i for i, f in enumerate(filas)
                    if f and len(f) > 1 and f[1] and norm(f[1]).startswith("ene - mar"))
    n_per = contar_periodos(filas, fila_hdr)
    periodos = periodos_desde(ANCLA_INFORMALIDAD, n_per)

    datos = {}
    for f in filas[fila_hdr + 1:]:
        if not f or f[0] is None:
            continue
        if norm(f[0]).startswith("fuente"):
            break
        nombre = titulo_ciudad(f[0])
        datos[nombre] = {"prop_informalidad":
                         [num(f[c]) if c < len(f) else None for c in range(1, 1 + n_per)]}

    # 2) niveles ocupados / formal / informal por ciudad
    filas2 = cargar_hoja(ruta, "Ciudades")
    fila_hdr2 = next(i for i, f in enumerate(filas2)
                     if f and any(x and norm(x).startswith("ene - mar") for x in f[:4]))
    col_ini2 = next(c for c, x in enumerate(filas2[fila_hdr2])
                    if x and norm(x).startswith("ene - mar"))
    nombre_actual = None
    for f in filas2[fila_hdr2 + 1:]:
        if not f:
            continue
        if f[0] is not None and str(f[0]).strip():
            if norm(f[0]).startswith("fuente"):
                break
            nombre_actual = titulo_ciudad(f[0])
        if nombre_actual is None or len(f) < 2 or f[1] is None:
            continue
        etq = norm(f[1])
        clave = ("ocupados" if etq.startswith("poblacion ocupada")
                 else "formal" if etq.startswith("formal")
                 else "informal" if etq.startswith("informal") else None)
        if clave is None:
            continue
        datos.setdefault(nombre_actual, {})[clave] = [
            num(f[c]) if c < len(f) else None
            for c in range(col_ini2, col_ini2 + n_per)]

    print(f"  informalidad : {len(datos)} ciudades x {n_per} trimestres moviles "
          f"({periodos[0][1]} -> {periodos[-1][1]})")
    return periodos, datos


# ---------------------------------------------------------------------------
# Promedio del anio
# ---------------------------------------------------------------------------

NIVELES = ("pob_total", "pet", "ft", "ocupados", "desocupados", "ffft",
           "subocupados", "formal", "informal", "pet_joven", "pet_total")

# El promedio anual se define aqui: los cuatro trimestres moviles de cierre.
TRIMESTRES_CIERRE = (3, 6, 9, 12)


def _promediar(datos_tm, por_anio, anios, limite):
    """
    Promedia por anio los trimestres de cierre. `limite` dice cuantos
    trimestres usar de cada anio (para poder comparar tramos equivalentes).
    """
    salida = {}
    for ciudad, ind in datos_tm.items():
        d = {}
        for clave in NIVELES:
            serie = ind.get(clave)
            if not serie:
                continue
            fila = []
            for a in anios:
                idxs = por_anio.get(a, [])[:limite.get(a, 4)]
                vals = [serie[i] for i in idxs if i < len(serie) and serie[i] is not None]
                # Un promedio con trimestres faltantes no es el promedio del
                # anio: es el de los meses que sobrevivieron. Mejor dejarlo vacio.
                fila.append(round(sum(vals) / len(vals), 2)
                            if vals and len(vals) == len(idxs) else None)
            d[clave] = fila

        def tasa(numer, denom):
            if not numer or not denom:
                return None
            return [round(x / y * 100, 2) if (x is not None and y) else None
                    for x, y in zip(numer, denom)]

        pet, ft, oc = d.get("pet"), d.get("ft"), d.get("ocupados")
        for clave, val in (("tgp", tasa(ft, pet)), ("to", tasa(oc, pet)),
                           ("td", tasa(d.get("desocupados"), ft)),
                           ("ts", tasa(d.get("subocupados"), ft)),
                           ("pct_pet", tasa(pet, d.get("pob_total"))),
                           ("prop_informalidad", tasa(d.get("informal"), oc))):
            if val:
                d[clave] = val

        # La juventud tiene su propio denominador de poblacion en edad de trabajar
        if d.get("pet_joven"):
            for clave, val in (("tgp", tasa(ft, d["pet_joven"])),
                               ("to", tasa(oc, d["pet_joven"]))):
                if val:
                    d[clave] = val

        salida[ciudad] = d
    return salida


def cierres_por_anio(periodos_tm):
    """Posicion de los trimestres moviles Ene-Mar, Abr-Jun, Jul-Sep y Oct-Dic."""
    por_anio = {}
    for i, (_, _, a, m) in enumerate(periodos_tm):
        if m in TRIMESTRES_CIERRE:
            por_anio.setdefault(a, []).append(i)
    return por_anio


def promedio_anual(periodos_tm, datos_tm):
    """
    Promedio del anio segun la definicion de la CCC: los cuatro trimestres
    moviles que NO se traslapan -- Ene-Mar, Abr-Jun, Jul-Sep, Oct-Dic -- de
    modo que cada mes entra exactamente una vez.

    Los NIVELES se promedian; las TASAS se RECALCULAN desde esos niveles,
    porque promediar tasas sesga el resultado hacia los trimestres con
    menos poblacion.

    Devuelve tambien una serie de REFERENCIA: para cada anio, el valor del
    anio anterior medido con el mismo numero de trimestres. Asi un anio en
    curso (2026 hasta junio) se compara contra el mismo tramo del anio
    pasado y no contra un anio entero.
    """
    por_anio = cierres_por_anio(periodos_tm)
    anios = sorted(por_anio)
    n_de = {a: len(por_anio[a]) for a in anios}

    periodos_an = [(str(a), etiqueta_anual(a, n_de[a]), a, n_de[a]) for a in anios]
    serie = _promediar(datos_tm, por_anio, anios, n_de)

    # Referencia: anio A-1 recortado al mismo numero de trimestres que A
    anios_ref = [a - 1 for a in anios]
    lim_ref = {a - 1: n_de[a] for a in anios}
    ref_bruta = _promediar(datos_tm, por_anio, anios_ref, lim_ref)

    parciales = [a for a in anios if n_de[a] < 4]
    etq_ref = [(f"{a-1} (Ene-{MESES[n_de[a]*3-1]})" if n_de[a] < 4 else str(a - 1))
               if (a - 1) in por_anio else None for a in anios]
    return periodos_an, serie, ref_bruta, parciales, etq_ref


# ---------------------------------------------------------------------------
# Ensamblaje
# ---------------------------------------------------------------------------

def indexar(periodos, clave_extra="meses"):
    return {"codigos": [p[0] for p in periodos],
            "etiquetas": [p[1] for p in periodos],
            "anios": [p[2] for p in periodos],
            clave_extra: [p[3] for p in periodos]}


def alinear(datos_origen, codigos_origen, codigos_destino):
    """Reindexa una serie a la grilla de periodos de destino (rellena con None)."""
    pos = {c: i for i, c in enumerate(codigos_origen)}
    salida = {}
    for ciudad, ind in datos_origen.items():
        d = {}
        for clave, serie in ind.items():
            d[clave] = [serie[pos[c]] if c in pos and pos[c] < len(serie) else None
                        for c in codigos_destino]
        salida[ciudad] = d
    return salida


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--datos", default=str(RAIZ / "datos"))
    ap.add_argument("--salida", default=str(RAIZ / "docs" / "datos.json"))
    args = ap.parse_args()

    carpeta = Path(args.datos)
    print(f"\nLeyendo anexos GEIH de {carpeta}")
    anexos = localizar_anexos(carpeta)
    print()

    per_tm, gen_tm = parse_general(anexos["general"])
    tn = parse_total_nacional(anexos["general"], len(per_tm))
    if tn:
        gen_tm["Total nacional"] = tn
        print("  total nacional: agregado al modulo general")
    _, sexo_tm = parse_sexo(anexos["sexo"])
    _, jov_tm = parse_juventud(anexos["juventud"])
    per_inf, inf_tm = parse_informalidad(anexos["informalidad"])
    cod_tm = [p[0] for p in per_tm]

    # informalidad reindexada a la grilla completa de trimestres moviles
    inf_alin = alinear(inf_tm, [p[0] for p in per_inf], cod_tm)

    # --- promedio del anio ---
    per_an, gen_an, gen_ref, parciales, etq_ref = promedio_anual(per_tm, gen_tm)
    _, hom_an, hom_ref, _, _ = promedio_anual(per_tm, sexo_tm["hombres"])
    _, muj_an, muj_ref, _, _ = promedio_anual(per_tm, sexo_tm["mujeres"])
    _, jov_an, jov_ref, _, _ = promedio_anual(per_tm, jov_tm)
    _, inf_an, inf_ref, _, _ = promedio_anual(per_tm, inf_alin)
    print(f"  promedio anual: {len(per_an)} anios "
          f"({per_an[0][1]} -> {per_an[-1][1]})"
          + (f"; parcial: {', '.join(map(str, parciales))}" if parciales else ""))

    universo = set(gen_tm) | set(jov_tm) | set(inf_alin) | set(sexo_tm["mujeres"])
    ciudades = [c for c in ORDEN_CIUDADES if c in universo]
    ciudades += sorted(c for c in universo if c not in ciudades)

    bonitos = {c: NOMBRES_BONITOS.get(c, c) for c in ciudades}

    salida = {
        "nombres": bonitos,
        "meta": {
            "generado": date.today().isoformat(),
            "fuente": "DANE - Gran Encuesta Integrada de Hogares (GEIH)",
            "elaboracion": "Camara de Comercio de Cali",
            "ciudad_foco": CIUDAD_FOCO,
            "ultimo_tm": per_tm[-1][1],
            "ultimo_anio": per_an[-1][1],
            "anios_parciales": parciales,
            "archivos": urls_de_origen(anexos),
            "nota_anual": (
                "El promedio del anio se calcula con los cuatro trimestres moviles que no "
                "se traslapan -- Ene-Mar, Abr-Jun, Jul-Sep y Oct-Dic -- de modo que cada mes "
                "entra una sola vez. Los niveles de poblacion se promedian y las tasas se "
                "recalculan a partir de esos niveles."
            ),
            "inicio_informalidad": per_inf[0][1],
        },
        "ciudades": ciudades,
        "periodos": {"tm": indexar(per_tm),
                     "an": {**indexar(per_an, "trimestres"), "referencia": etq_ref}},
        "series": {
            "tm": {
                "general": gen_tm,
                "hombres": sexo_tm["hombres"],
                "mujeres": sexo_tm["mujeres"],
                "juventud": jov_tm,
                "informalidad": inf_alin,
            },
            "an": {
                "general": gen_an,
                "hombres": hom_an,
                "mujeres": muj_an,
                "juventud": jov_an,
                "informalidad": inf_an,
            },
        },
        # Mismo tramo del anio anterior: hace comparable un anio en curso
        "referencia_anual": {
            "general": gen_ref, "hombres": hom_ref, "mujeres": muj_ref,
            "juventud": jov_ref, "informalidad": inf_ref,
        },
    }

    destino = Path(args.salida)
    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_text(json.dumps(salida, ensure_ascii=False, separators=(",", ":")),
                       encoding="utf-8")
    mb = destino.stat().st_size / 1e6
    print(f"\n[listo] {destino}  ({mb:.2f} MB)")
    print(f"        ultimo trimestre movil: {per_tm[-1][1]}")
    print(f"        ultimo anio           : {per_an[-1][1]}\n")


if __name__ == "__main__":
    main()
